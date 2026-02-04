#!/usr/bin/env python3
"""
Add building records for parcels that have building_sf data but no building record.
This fixes the issue where parcels with buildings are incorrectly shown as "land only".
"""

import os
import sys
import csv
import psycopg2
from pathlib import Path

# Database connection
DATABASE_URL = 'postgresql://postgres.qrgkwcofdwkodxanaubr:$ealTheDeal51!@aws-0-us-west-2.pooler.supabase.com:5432/postgres'

def main():
    print("=" * 60)
    print("Add Missing Building Records")
    print("=" * 60)

    # Connect to database
    print("\n1. Connecting to database...")
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # Get all parcels
    print("2. Fetching all parcels...")
    cur.execute("SELECT apn, situs_address, city, zip FROM parcel")
    parcels = {row[0]: {'address': row[1], 'city': row[2], 'zip': row[3]} for row in cur.fetchall()}
    print(f"   Found {len(parcels)} parcels")

    # Get existing building records
    print("3. Fetching existing buildings...")
    cur.execute("SELECT DISTINCT parcel_apn FROM building")
    existing_building_apns = {row[0] for row in cur.fetchall()}
    print(f"   Found {len(existing_building_apns)} parcels with buildings")

    # Find parcels without buildings
    parcels_without_buildings = set(parcels.keys()) - existing_building_apns
    print(f"   Parcels without building records: {len(parcels_without_buildings)}")

    # Load source data for building_sqft
    print("\n4. Loading source data for building_sqft...")

    # Map APN -> building data
    apn_to_building_data = {}

    # Load from Downloads/Parcels.csv
    downloads_file = Path("C:/Users/User/Downloads/Parcels.csv")
    if downloads_file.exists():
        with open(downloads_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                apn = row.get('APN', '').strip()
                building_sf = row.get('BUILDING_SQFT', '')
                year_built = row.get('YR_BLT', '')
                use_type = row.get('USE_CODE_STD_CTGR_DESC', '')

                try:
                    building_sf = float(building_sf) if building_sf else 0
                except:
                    building_sf = 0

                try:
                    year_built = int(float(year_built)) if year_built else None
                except:
                    year_built = None

                if apn and building_sf > 0:
                    apn_to_building_data[apn] = {
                        'building_sf': int(building_sf),
                        'year_built': year_built if year_built and year_built > 1800 else None,
                        'property_type': use_type or 'INDUSTRIAL'
                    }
        print(f"   Loaded {len(apn_to_building_data)} records from Downloads/Parcels.csv")

    # Also try to load from Excel files
    try:
        import openpyxl
        backend_dir = Path("C:/Users/User/industrial-tracker/backend")
        excel_files = list(backend_dir.glob("*.xlsx"))

        for excel_file in excel_files:
            try:
                wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]

                apn_idx = headers.index('APN') if 'APN' in headers else None
                bldg_sf_idx = headers.index('BUILDING_SQFT') if 'BUILDING_SQFT' in headers else None
                yr_built_idx = headers.index('YR_BLT') if 'YR_BLT' in headers else None
                use_idx = headers.index('USE_CODE_STD_CTGR_DESC') if 'USE_CODE_STD_CTGR_DESC' in headers else None

                if apn_idx is not None and bldg_sf_idx is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        apn = str(row[apn_idx] or '').strip()
                        building_sf = row[bldg_sf_idx] or 0

                        try:
                            building_sf = float(building_sf)
                        except:
                            building_sf = 0

                        if apn and building_sf > 0 and apn not in apn_to_building_data:
                            year_built = None
                            if yr_built_idx is not None:
                                try:
                                    year_built = int(row[yr_built_idx]) if row[yr_built_idx] else None
                                except:
                                    pass

                            use_type = row[use_idx] if use_idx is not None else None

                            apn_to_building_data[apn] = {
                                'building_sf': int(building_sf),
                                'year_built': year_built if year_built and year_built > 1800 else None,
                                'property_type': use_type or 'INDUSTRIAL'
                            }

                wb.close()
            except Exception as e:
                pass

        print(f"   Total building data records: {len(apn_to_building_data)}")
    except ImportError:
        print("   openpyxl not installed, skipping Excel files")

    # Create normalized lookup for building data
    normalized_building_data = {}
    for apn, data in apn_to_building_data.items():
        normalized = apn.replace('-', '')
        normalized_building_data[normalized] = data

    # Find parcels that need building records
    print("\n5. Finding parcels that need building records...")
    buildings_to_add = []

    for apn in parcels_without_buildings:
        # Normalize APN for lookup (remove dashes)
        normalized_apn = apn.replace('-', '')

        building_data = normalized_building_data.get(normalized_apn)

        if building_data:
            parcel_info = parcels[apn]
            buildings_to_add.append({
                'parcel_apn': apn,
                'address': parcel_info['address'],
                'city': parcel_info['city'],
                'state': 'CA',
                'zip': parcel_info['zip'],
                'building_sf': building_data['building_sf'],
                'year_built': building_data['year_built'],
                'property_type': building_data['property_type'],
                'property_subtype': None
            })

    print(f"   Found {len(buildings_to_add)} parcels with building data but no building record")

    if not buildings_to_add:
        print("\nNo buildings to add!")
        return

    # Insert missing building records
    print(f"\n6. Inserting {len(buildings_to_add)} building records...")

    # Building table columns: id, parcel_apn, building_name, building_sf, year_built,
    # construction_type, office_stories, sprinklers, notes, created_at, updated_at

    inserted = 0
    errors = 0

    for building in buildings_to_add:
        try:
            cur.execute("""
                INSERT INTO building (parcel_apn, building_sf, year_built)
                VALUES (%s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (
                building['parcel_apn'],
                building['building_sf'],
                building['year_built']
            ))
            inserted += 1
        except Exception as e:
            errors += 1
            if errors <= 3:
                print(f"   Error: {e}")

        if inserted % 500 == 0:
            conn.commit()
            print(f"   Progress: {inserted}/{len(buildings_to_add)}")

    conn.commit()

    print(f"\n7. Results:")
    print(f"   Inserted: {inserted}")
    print(f"   Errors: {errors}")

    # Verify final counts
    print("\n8. Final verification...")
    cur.execute("SELECT COUNT(*) FROM building")
    building_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(DISTINCT parcel_apn) FROM building")
    parcels_with_buildings = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM parcel")
    parcel_count = cur.fetchone()[0]

    print(f"   Total parcels: {parcel_count}")
    print(f"   Total buildings: {building_count}")
    print(f"   Parcels with buildings: {parcels_with_buildings}")
    print(f"   Parcels without buildings (land): {parcel_count - parcels_with_buildings}")

    cur.close()
    conn.close()

    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)


if __name__ == '__main__':
    main()
