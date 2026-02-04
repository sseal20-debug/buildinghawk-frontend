#!/usr/bin/env python3
"""
Import comp data from Excel files into the PostgreSQL database.
Handles various Excel formats found in the Dropbox comp folders.
Extracts ALL property data to fill comp fields comprehensively.
"""

import os
import re
import sys
import json
from datetime import datetime
from pathlib import Path
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# Database connection - uses Supabase
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres.qrgkwcofdwkodxanaubr:$ealTheDeal51!@aws-0-us-west-2.pooler.supabase.com:5432/postgres')

# Dropbox comp folders
COMP_FOLDER = r"C:\Users\User\Seal Industrial Dropbox\1    SOLD or LEASED_sealhudmag\1     NOC LEASE AND SALE COMPS"


def parse_sf(value):
    """Parse square footage from various formats."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value) if value > 0 else None
    if isinstance(value, str):
        cleaned = re.sub(r'[^\d.]', '', value)
        # Handle cases like '.' or empty after cleaning
        if cleaned and cleaned != '.' and cleaned.replace('.', ''):
            try:
                return int(float(cleaned))
            except ValueError:
                return None
    return None


def parse_price(value):
    """Parse price from various formats."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    if isinstance(value, str):
        cleaned = re.sub(r'[^\d.]', '', value)
        # Handle cases like '.' or empty after cleaning
        if cleaned and cleaned != '.' and cleaned.replace('.', ''):
            try:
                return float(cleaned)
            except ValueError:
                return None
    return None


def parse_date(value):
    """Parse date from various formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        # Try various date formats
        formats = [
            '%b-%y',      # Mar-21
            '%B %Y',      # March 2021
            '%b %Y',      # Mar 2021
            '%m/%d/%Y',   # 03/15/2021
            '%Y-%m-%d',   # 2021-03-15
            '%b %y',      # Mar 21
            '%m-%y',      # 03-21
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def parse_rate(value):
    """Parse rental rate or percentage."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        match = re.search(r'([\d.]+)', value)
        if match:
            return float(match.group(1))
    return None


def parse_height(value):
    """Parse clear height from formats like "26'" or "22'-24'"."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        match = re.search(r'(\d+)', value)
        if match:
            return float(match.group(1))
    return None


def parse_doors(value):
    """Parse door count."""
    if value is None or value == 'None' or value == 'Not Available':
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        match = re.search(r'(\d+)', value)
        if match:
            return int(match.group(1))
    return None


def parse_percentage(value):
    """Parse percentage from formats like '3% annually' or 0.03."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # If value is less than 1, assume it's a decimal (0.03 = 3%)
        if 0 < value < 1:
            return value * 100
        return float(value)
    if isinstance(value, str):
        match = re.search(r'([\d.]+)', value)
        if match:
            val = float(match.group(1))
            if 0 < val < 1:
                return val * 100
            return val
    return None


def extract_city_from_address(address):
    """Try to extract city from address string."""
    if not address:
        return None, None

    address = str(address)

    # Handle formats like "1651 E Saint Andrew Pl\nSanta Ana"
    if '\n' in address:
        parts = address.split('\n')
        street = parts[0].strip()
        city_part = parts[1].strip() if len(parts) > 1 else None
        # Remove "Industrial" suffix
        if city_part:
            city = city_part.replace(' Industrial', '').strip()
            return street, city

    # Handle comma-separated format "123 Main St, Anaheim"
    if ',' in address:
        parts = address.split(',')
        street = parts[0].strip()
        city = parts[1].strip() if len(parts) > 1 else None
        return street, city

    return address, None


def normalize_lease_structure(value):
    """Normalize lease structure to database enum values."""
    if not value:
        return None
    value_upper = str(value).upper().strip()
    mapping = {
        'NNN': 'nnn',
        'TRIPLE NET': 'nnn',
        'N': 'nnn',
        'G': 'gross',
        'GROSS': 'gross',
        'MG': 'modified_gross',
        'MODIFIED GROSS': 'modified_gross',
        'IG': 'industrial_gross',
        'INDUSTRIAL GROSS': 'industrial_gross',
        'FSG': 'fsg',
        'FULL SERVICE GROSS': 'fsg',
    }
    return mapping.get(value_upper, 'nnn')


def import_lease_comps_horizontal(wb, sheet_name, conn, source_file):
    """Import lease comps from horizontal format (properties in columns)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return 0

    # Build comprehensive field mapping from row labels
    field_map = {}
    for row_idx, row in enumerate(rows):
        if row[0]:
            label = str(row[0]).lower().strip()

            # Address fields
            if 'address' in label or 'street' in label:
                field_map['address'] = row_idx

            # Size fields
            elif 'unit size' in label:
                field_map['leased_sf'] = row_idx
            elif 'project size' in label or 'building' in label and ('sf' in label or 'area' in label):
                field_map['building_sf'] = row_idx
            elif 'office' in label and 'sf' in label:
                field_map['office_sf'] = row_idx
            elif 'land' in label and ('ratio' in label or 'building' in label):
                field_map['land_ratio'] = row_idx

            # Building specs
            elif 'year built' in label:
                field_map['year_built'] = row_idx
            elif 'ground-level' in label or 'ground level' in label:
                field_map['gl_doors'] = row_idx
            elif 'dock' in label and ('high' in label or 'door' in label):
                field_map['dock_doors'] = row_idx
            elif 'clear height' in label:
                field_map['clear_height'] = row_idx
            elif 'office' in label and '%' in label:
                field_map['office_pct'] = row_idx
            elif 'parking' in label:
                field_map['parking'] = row_idx
            elif 'construction' in label:
                field_map['construction'] = row_idx
            elif 'occupancy' in label and 'sale' not in label:
                field_map['occupancy'] = row_idx

            # Lease terms
            elif 'lease date' in label or 'rental data' in label:
                field_map['lease_date'] = row_idx
            elif 'lease term' in label:
                field_map['term'] = row_idx
            elif 'lease type' in label:
                field_map['lease_type'] = row_idx
            elif 'escalation' in label or 'adjustment' in label:
                field_map['escalations'] = row_idx
            elif 'free rent' in label:
                field_map['free_rent'] = row_idx
            elif 'ti allowance' in label or ('ti' in label.split() and 'ti' == label.split()[0]):
                field_map['ti'] = row_idx

            # Rent fields
            elif label == 'rent' or (('rent' in label or 'rate' in label) and 'free' not in label and 'escalation' not in label and 'adjustment' not in label):
                if 'rent' not in field_map:
                    field_map['rent'] = row_idx
            elif 'nnn' in label or 'cam' in label:
                field_map['nnn_charges'] = row_idx
            elif 'total' in label and ('mo' in label or 'month' in label or 'p/mo' in label):
                field_map['total_rent'] = row_idx

            # Notes
            elif 'notes' in label or 'comment' in label:
                field_map['notes'] = row_idx

    # Extract comps from columns
    comps = []
    num_cols = len(rows[0]) if rows else 0

    for col_idx in range(1, num_cols):
        address_row = field_map.get('address', 0)
        raw_address = rows[address_row][col_idx] if address_row < len(rows) else None

        if not raw_address or raw_address is None:
            continue

        street, city = extract_city_from_address(str(raw_address))
        if not street:
            continue

        if not city:
            city = 'Anaheim'

        # Get lease date, default to estimated date
        lease_date = None
        if 'lease_date' in field_map and field_map['lease_date'] < len(rows):
            lease_date = parse_date(rows[field_map['lease_date']][col_idx])
        if not lease_date:
            lease_date = '2021-01-01'  # Default

        # Get term in years, convert to months
        term_months = None
        if 'term' in field_map and field_map['term'] < len(rows):
            term_years = parse_rate(rows[field_map['term']][col_idx])
            if term_years:
                term_months = int(term_years * 12) if term_years < 100 else int(term_years)

        # Get office SF
        office_sf = None
        if 'office_sf' in field_map and field_map['office_sf'] < len(rows):
            office_sf = parse_sf(rows[field_map['office_sf']][col_idx])
        elif 'office_pct' in field_map and field_map['office_pct'] < len(rows):
            office_pct = parse_rate(rows[field_map['office_pct']][col_idx])
            leased_sf = parse_sf(rows[field_map['leased_sf']][col_idx]) if 'leased_sf' in field_map else None
            if office_pct and leased_sf:
                office_sf = int(leased_sf * (office_pct if office_pct < 1 else office_pct / 100))

        comp = {
            'property_address': street,
            'city': city,
            'state': 'CA',
            'submarket': 'North Orange County',
            'leased_sf': parse_sf(rows[field_map['leased_sf']][col_idx]) if 'leased_sf' in field_map and field_map['leased_sf'] < len(rows) else None,
            'building_sf': parse_sf(rows[field_map['building_sf']][col_idx]) if 'building_sf' in field_map and field_map['building_sf'] < len(rows) else None,
            'office_sf': office_sf,
            'year_built': parse_sf(rows[field_map['year_built']][col_idx]) if 'year_built' in field_map and field_map['year_built'] < len(rows) else None,
            'gl_doors': parse_doors(rows[field_map['gl_doors']][col_idx]) if 'gl_doors' in field_map and field_map['gl_doors'] < len(rows) else None,
            'dock_doors': parse_doors(rows[field_map['dock_doors']][col_idx]) if 'dock_doors' in field_map and field_map['dock_doors'] < len(rows) else None,
            'clear_height_ft': parse_height(rows[field_map['clear_height']][col_idx]) if 'clear_height' in field_map and field_map['clear_height'] < len(rows) else None,
            'lease_date': lease_date,
            'lease_term_months': term_months,
            'lease_structure': normalize_lease_structure(rows[field_map['lease_type']][col_idx]) if 'lease_type' in field_map and field_map['lease_type'] < len(rows) else 'nnn',
            'starting_rent_psf': parse_rate(rows[field_map['rent']][col_idx]) if 'rent' in field_map and field_map['rent'] < len(rows) else None,
            'nnn_expenses_psf': parse_rate(rows[field_map['nnn_charges']][col_idx]) if 'nnn_charges' in field_map and field_map['nnn_charges'] < len(rows) else None,
            'annual_increases': parse_percentage(rows[field_map['escalations']][col_idx]) if 'escalations' in field_map and field_map['escalations'] < len(rows) else None,
            'free_rent_months': parse_doors(rows[field_map['free_rent']][col_idx]) if 'free_rent' in field_map and field_map['free_rent'] < len(rows) else None,
            'ti_allowance_psf': parse_rate(rows[field_map['ti']][col_idx]) if 'ti' in field_map and field_map['ti'] < len(rows) else None,
            'notes': str(rows[field_map['notes']][col_idx]) if 'notes' in field_map and field_map['notes'] < len(rows) and rows[field_map['notes']][col_idx] else None,
        }

        # Use leased_sf as default if not set but building_sf is
        if not comp['leased_sf'] and comp['building_sf']:
            comp['leased_sf'] = comp['building_sf']

        # Only add if we have meaningful data
        if comp['leased_sf'] or comp['starting_rent_psf']:
            comps.append(comp)

    if comps:
        insert_lease_comps(conn, comps, source_file)

    return len(comps)


def import_sale_comps_horizontal(wb, sheet_name, conn, source_file):
    """Import sale comps from horizontal format (properties in columns)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return 0

    # Build comprehensive field mapping
    field_map = {}
    for row_idx, row in enumerate(rows):
        if row[0]:
            label = str(row[0]).lower().strip()

            # Address
            if 'address' in label or 'street' in label:
                field_map['address'] = row_idx
            elif 'apn' in label:
                field_map['apn'] = row_idx

            # Size
            elif 'building area' in label or 'building sf' in label:
                field_map['sf'] = row_idx
            elif 'land area' in label:
                field_map['land_sf'] = row_idx
            elif 'land' in label and ('ratio' in label or 'bldg' in label):
                field_map['land_ratio'] = row_idx
            elif 'office' in label and ('percent' in label or '%' in label):
                field_map['office_pct'] = row_idx

            # Building specs
            elif 'year built' in label or 'yr blt' in label:
                field_map['year_built'] = row_idx
            elif 'ground-level' in label or 'ground level' in label:
                field_map['gl_doors'] = row_idx
            elif 'dock' in label:
                field_map['dock_doors'] = row_idx
            elif 'clear height' in label:
                field_map['clear_height'] = row_idx
            elif 'construction' in label:
                field_map['construction'] = row_idx
            elif 'parking' in label:
                field_map['parking'] = row_idx

            # Sale info
            elif 'sale' in label and 'date' in label:
                field_map['sale_date'] = row_idx
            elif 'sale price' in label and 'sf' not in label and '/' not in label:
                field_map['sale_price'] = row_idx
            elif ('price' in label and ('sf' in label or '/sf' in label)) or 'price / sf' in label:
                field_map['price_psf'] = row_idx
            elif 'oar' in label or 'cap rate' in label:
                field_map['cap_rate'] = row_idx
            elif 'noi' in label:
                field_map['noi'] = row_idx
            elif 'occupancy' in label:
                field_map['occupancy'] = row_idx

    # Extract comps
    comps = []
    num_cols = len(rows[0]) if rows else 0

    for col_idx in range(1, num_cols):
        address_row = field_map.get('address', 0)
        raw_address = rows[address_row][col_idx] if address_row < len(rows) else None

        if not raw_address or raw_address is None:
            continue

        street, city = extract_city_from_address(str(raw_address))
        if not street:
            continue

        if not city:
            city = 'Orange'

        sale_price = parse_price(rows[field_map['sale_price']][col_idx]) if 'sale_price' in field_map and field_map['sale_price'] < len(rows) else None
        price_psf = parse_rate(rows[field_map['price_psf']][col_idx]) if 'price_psf' in field_map and field_map['price_psf'] < len(rows) else None
        building_sf = parse_sf(rows[field_map['sf']][col_idx]) if 'sf' in field_map and field_map['sf'] < len(rows) else None

        # Calculate price_psf if we have sale_price and sf
        if sale_price and building_sf and not price_psf:
            price_psf = round(sale_price / building_sf, 2)

        # Get sale date
        sale_date = None
        if 'sale_date' in field_map and field_map['sale_date'] < len(rows):
            sale_date = parse_date(rows[field_map['sale_date']][col_idx])
        if not sale_date:
            sale_date = '2021-01-01'

        # Get land area
        land_sf = parse_sf(rows[field_map['land_sf']][col_idx]) if 'land_sf' in field_map and field_map['land_sf'] < len(rows) else None

        # Get cap rate (convert from decimal if needed)
        cap_rate = None
        if 'cap_rate' in field_map and field_map['cap_rate'] < len(rows):
            raw_cap = rows[field_map['cap_rate']][col_idx]
            if raw_cap and raw_cap not in ['#VALUE!', '---', '-']:
                cap_rate = parse_rate(raw_cap)
                if cap_rate and cap_rate < 1:
                    cap_rate = cap_rate * 100  # Convert from decimal

        # Get NOI
        noi = parse_price(rows[field_map['noi']][col_idx]) if 'noi' in field_map and field_map['noi'] < len(rows) else None

        # Get occupancy
        occupancy = None
        if 'occupancy' in field_map and field_map['occupancy'] < len(rows):
            occ_val = rows[field_map['occupancy']][col_idx]
            if occ_val:
                occupancy = parse_rate(occ_val)
                if occupancy and occupancy <= 1:
                    occupancy = occupancy * 100

        comp = {
            'property_address': street,
            'city': city,
            'state': 'CA',
            'submarket': 'North Orange County',
            'building_sf': building_sf,
            'land_sf': land_sf,
            'land_acres': round(land_sf / 43560, 4) if land_sf else None,
            'year_built': parse_sf(rows[field_map['year_built']][col_idx]) if 'year_built' in field_map and field_map['year_built'] < len(rows) else None,
            'gl_doors': parse_doors(rows[field_map['gl_doors']][col_idx]) if 'gl_doors' in field_map and field_map['gl_doors'] < len(rows) else None,
            'dock_doors': parse_doors(rows[field_map['dock_doors']][col_idx]) if 'dock_doors' in field_map and field_map['dock_doors'] < len(rows) else None,
            'clear_height_ft': parse_height(rows[field_map['clear_height']][col_idx]) if 'clear_height' in field_map and field_map['clear_height'] < len(rows) else None,
            'sale_date': sale_date,
            'sale_price': sale_price,
            'price_psf': price_psf,
            'cap_rate': cap_rate,
            'noi': noi,
            'occupancy_pct': occupancy,
        }

        if comp['building_sf'] or comp['sale_price']:
            comps.append(comp)

    if comps:
        insert_sale_comps(conn, comps, source_file)

    return len(comps)


def import_sale_comps_vertical(wb, sheet_name, conn, source_file):
    """Import sale comps from vertical format (headers in first row, data in rows)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return 0

    # First row is headers
    headers = [str(h).lower().strip() if h else '' for h in rows[0]]

    # Build column mapping
    col_map = {}
    for idx, header in enumerate(headers):
        if 'address' in header:
            col_map['address'] = idx
        elif header == 'size' or ('building' in header and 'sf' in header) or 'sf' in header.split():
            col_map['sf'] = idx
        elif 'yr blt' in header or 'year' in header:
            col_map['year_built'] = idx
        elif 'sale date' in header:
            col_map['sale_date'] = idx
        elif 'sale price' in header and 'sf' not in header:
            col_map['sale_price'] = idx
        elif 'price/sf' in header or 'price psf' in header or '/sf' in header:
            col_map['price_psf'] = idx
        elif 'cap rate' in header:
            col_map['cap_rate'] = idx
        elif 'bus park' in header or 'assoc' in header:
            col_map['bus_park'] = idx
        elif 'type' in header:
            col_map['type'] = idx

    comps = []
    for row in rows[1:]:
        if not row:
            continue

        address_idx = col_map.get('address', 0)
        if address_idx >= len(row) or not row[address_idx]:
            continue

        address = str(row[address_idx])
        city = 'Orange'  # Default

        sf = parse_sf(row[col_map['sf']]) if 'sf' in col_map and col_map['sf'] < len(row) else None

        year_str = row[col_map.get('year_built')] if 'year_built' in col_map and col_map['year_built'] < len(row) else None
        year_built = None
        if year_str:
            match = re.search(r'(\d{4})', str(year_str))
            if match:
                year_built = int(match.group(1))

        sale_price = parse_price(row[col_map['sale_price']]) if 'sale_price' in col_map and col_map['sale_price'] < len(row) else None
        price_psf = parse_rate(row[col_map['price_psf']]) if 'price_psf' in col_map and col_map['price_psf'] < len(row) else None

        if sale_price and sf and not price_psf:
            price_psf = round(sale_price / sf, 2)

        sale_date = None
        if 'sale_date' in col_map and col_map['sale_date'] < len(row):
            date_val = row[col_map['sale_date']]
            if date_val:
                sale_date = parse_date(date_val)
        if not sale_date:
            sale_date = '2023-01-01'

        cap_rate = None
        if 'cap_rate' in col_map and col_map['cap_rate'] < len(row):
            raw_cap = row[col_map['cap_rate']]
            if raw_cap:
                cap_rate = parse_rate(raw_cap)

        comp = {
            'property_address': address,
            'city': city,
            'state': 'CA',
            'submarket': 'North Orange County',
            'building_sf': sf,
            'land_sf': None,
            'land_acres': None,
            'year_built': year_built,
            'dock_doors': None,
            'gl_doors': None,
            'clear_height_ft': None,
            'sale_date': sale_date,
            'sale_price': sale_price,
            'price_psf': price_psf,
            'cap_rate': cap_rate,
            'noi': None,
            'occupancy_pct': None,
        }

        if comp['building_sf'] or comp['sale_price']:
            comps.append(comp)

    if comps:
        insert_sale_comps(conn, comps, source_file)

    return len(comps)


def import_simple_lease_comps(wb, sheet_name, conn, source_file):
    """Import simple lease comps format (Address, City, SF, Rent, etc.)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return 0

    # Check if this is simple format with headers
    first_row = [str(h).lower() if h else '' for h in rows[0]]
    if 'address' in first_row[0]:
        # Headers in first row
        comps = []
        for row in rows[1:]:
            if not row or not row[0]:
                continue

            comp = {
                'property_address': str(row[0]) if row[0] else None,
                'city': str(row[1]) if len(row) > 1 and row[1] else 'Anaheim',
                'state': 'CA',
                'submarket': 'North Orange County',
                'building_sf': None,
                'leased_sf': parse_sf(row[2]) if len(row) > 2 else None,
                'office_sf': None,
                'year_built': None,
                'dock_doors': None,
                'gl_doors': None,
                'clear_height_ft': None,
                'starting_rent_psf': parse_rate(row[3]) if len(row) > 3 else None,
                'nnn_expenses_psf': None,
                'annual_increases': parse_percentage(row[4]) if len(row) > 4 else None,
                'free_rent_months': parse_doors(row[5]) if len(row) > 5 else None,
                'lease_term_months': None,
                'ti_allowance_psf': None,
                'notes': str(row[6]) if len(row) > 6 and row[6] else None,
                'lease_date': '2020-01-01',  # Default
                'lease_structure': 'nnn',
            }

            if comp['leased_sf'] or comp['starting_rent_psf']:
                comps.append(comp)

        if comps:
            insert_lease_comps(conn, comps, source_file)
        return len(comps)

    return 0


def clamp_numeric(value, max_val=999.99):
    """Clamp numeric values to fit database constraints."""
    if value is None:
        return None
    try:
        val = float(value)
        if val > max_val:
            return max_val
        if val < -max_val:
            return -max_val
        return val
    except (ValueError, TypeError):
        return None


def insert_lease_comps(conn, comps, source_file):
    """Insert lease comps into database matching actual schema."""
    cur = conn.cursor()
    inserted = 0

    for comp in comps:
        try:
            # Clamp numeric fields to fit database constraints
            comp['starting_rent_psf'] = clamp_numeric(comp.get('starting_rent_psf'))
            comp['nnn_expenses_psf'] = clamp_numeric(comp.get('nnn_expenses_psf'))
            comp['annual_increases'] = clamp_numeric(comp.get('annual_increases'), 99.99)
            comp['ti_allowance_psf'] = clamp_numeric(comp.get('ti_allowance_psf'))
            comp['clear_height_ft'] = clamp_numeric(comp.get('clear_height_ft'), 99.99)

            cur.execute("""
                INSERT INTO lease_comp (
                    property_address, city, state, submarket,
                    building_sf, leased_sf, office_sf, year_built,
                    dock_doors, gl_doors, clear_height_ft,
                    lease_date, lease_term_months, lease_structure,
                    starting_rent_psf, nnn_expenses_psf, annual_increases,
                    free_rent_months, ti_allowance_psf, notes, source
                ) VALUES (
                    %(property_address)s, %(city)s, %(state)s, %(submarket)s,
                    %(building_sf)s, %(leased_sf)s, %(office_sf)s, %(year_built)s,
                    %(dock_doors)s, %(gl_doors)s, %(clear_height_ft)s,
                    %(lease_date)s, %(lease_term_months)s, %(lease_structure)s,
                    %(starting_rent_psf)s, %(nnn_expenses_psf)s, %(annual_increases)s,
                    %(free_rent_months)s, %(ti_allowance_psf)s, %(notes)s, 'manual'
                )
            """, comp)
            inserted += 1
        except Exception as e:
            conn.rollback()  # Rollback failed transaction to allow next insert
            print(f"  Error inserting lease comp {comp.get('property_address')}: {e}")

    conn.commit()
    return inserted


def insert_sale_comps(conn, comps, source_file):
    """Insert sale comps into database matching actual schema."""
    cur = conn.cursor()
    inserted = 0

    for comp in comps:
        try:
            # Clamp numeric fields to fit database constraints
            comp['price_psf'] = clamp_numeric(comp.get('price_psf'))
            comp['cap_rate'] = clamp_numeric(comp.get('cap_rate'), 99.99)
            comp['clear_height_ft'] = clamp_numeric(comp.get('clear_height_ft'), 99.99)
            comp['occupancy_pct'] = clamp_numeric(comp.get('occupancy_pct'), 100)

            cur.execute("""
                INSERT INTO sale_comp (
                    property_address, city, state, submarket,
                    building_sf, land_sf, land_acres, year_built,
                    dock_doors, gl_doors, clear_height_ft,
                    sale_date, sale_price, price_psf, cap_rate, noi, occupancy_pct,
                    source
                ) VALUES (
                    %(property_address)s, %(city)s, %(state)s, %(submarket)s,
                    %(building_sf)s, %(land_sf)s, %(land_acres)s, %(year_built)s,
                    %(dock_doors)s, %(gl_doors)s, %(clear_height_ft)s,
                    %(sale_date)s, %(sale_price)s, %(price_psf)s, %(cap_rate)s, %(noi)s, %(occupancy_pct)s,
                    'manual'
                )
            """, comp)
            inserted += 1
        except Exception as e:
            conn.rollback()  # Rollback failed transaction to allow next insert
            print(f"  Error inserting sale comp {comp.get('property_address')}: {e}")

    conn.commit()
    return inserted


def process_excel_file(filepath, conn):
    """Process a single Excel file and import comps."""
    filename = os.path.basename(filepath)
    print(f"\nProcessing: {filename}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  Error opening file: {e}")
        return 0, 0

    lease_count = 0
    sale_count = 0

    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()

        if 'lease' in sheet_lower:
            count = import_lease_comps_horizontal(wb, sheet_name, conn, filename)
            if count == 0:
                count = import_simple_lease_comps(wb, sheet_name, conn, filename)
            lease_count += count
            print(f"  Lease sheet '{sheet_name}': {count} comps")

        elif 'sale' in sheet_lower or 'sold' in sheet_lower:
            ws = wb[sheet_name]
            first_row = list(ws.iter_rows(max_row=1, values_only=True))[0] if ws.max_row > 0 else []
            first_cell = str(first_row[0]).lower() if first_row and first_row[0] else ''

            if 'address' in first_cell:
                count = import_sale_comps_vertical(wb, sheet_name, conn, filename)
            else:
                count = import_sale_comps_horizontal(wb, sheet_name, conn, filename)

            sale_count += count
            print(f"  Sale sheet '{sheet_name}': {count} comps")

        elif 'nearby' in sheet_lower:
            count = import_sale_comps_vertical(wb, sheet_name, conn, filename)
            sale_count += count
            print(f"  Nearby Sale sheet '{sheet_name}': {count} comps")

        elif 'sheet1' in sheet_lower:
            # Check content to determine type
            ws = wb[sheet_name]
            all_text = ' '.join([str(cell) for row in ws.iter_rows(max_row=5, values_only=True) for cell in row if cell])

            if 'lease' in all_text.lower() or 'rent' in all_text.lower():
                count = import_simple_lease_comps(wb, sheet_name, conn, filename)
                lease_count += count
                if count:
                    print(f"  Generic sheet '{sheet_name}': {count} lease comps")

    return lease_count, sale_count


def main():
    """Main function to scan and import all comp Excel files."""
    print("=" * 60)
    print("Comp Data Import Script")
    print("Extracts ALL property data for comprehensive comp records")
    print("=" * 60)

    # Connect to database
    try:
        conn = psycopg2.connect(DATABASE_URL)
        print("Connected to database")
    except Exception as e:
        print(f"Database connection failed: {e}")
        sys.exit(1)

    # Find all Excel files
    excel_files = []
    for root, dirs, files in os.walk(COMP_FOLDER):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                file_lower = file.lower()
                if any(keyword in file_lower for keyword in ['comp', 'lease', 'sale', 'sold', 'orange county']):
                    excel_files.append(os.path.join(root, file))

    print(f"\nFound {len(excel_files)} potential comp files")

    total_lease = 0
    total_sale = 0

    for filepath in excel_files:
        lease_count, sale_count = process_excel_file(filepath, conn)
        total_lease += lease_count
        total_sale += sale_count

    print("\n" + "=" * 60)
    print(f"Import Complete!")
    print(f"  Total Lease Comps: {total_lease}")
    print(f"  Total Sale Comps: {total_sale}")
    print("=" * 60)

    conn.close()


if __name__ == '__main__':
    main()
